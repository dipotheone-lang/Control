from datetime import date, datetime

import pytest

from control.db import init_db
from control.report import HorizonItem, OpenItem, weekly_report

AS_OF = date(2026, 8, 13)


@pytest.fixture
def conn(tmp_path):
    c = init_db(tmp_path / "control.db")
    yield c
    c.close()


def _horizon():
    return [
        HorizonItem("STAT-VAT", 1, "VAT return", "accounts@ubcsis.com",
                    date(2026, 8, 28), "OPEN"),
        HorizonItem("TND-114", 2, "RFQ 114 submission", "donia@ubcsis.com",
                    date(2026, 8, 20), "ALERTED"),
        HorizonItem("ACC-SIEMENS", 2, "Siemens prequalification", "donia@ubcsis.com",
                    date(2026, 12, 1), "OPEN"),   # beyond 30 days -> excluded
    ]


def test_horizon_first_sorted_and_windowed(conn, tmp_path):
    r = weekly_report(conn, as_of=AS_OF, horizon=_horizon(), open_items=[],
                      open_decisions=[], control_root=tmp_path)
    body = r["body"]
    assert body.index("CLASS 1 & 2 HORIZON") < body.index("OPEN ITEMS")
    assert "RFQ 114" in body and "VAT return" in body
    assert "Siemens prequalification" not in body.split("2. OPEN")[0]
    # tender (20th, T-7) sorts before VAT (28th)
    assert body.index("RFQ 114") < body.index("VAT return")


def test_empty_horizon_states_the_gap(conn, tmp_path):
    r = weekly_report(conn, as_of=AS_OF, horizon=[], open_items=[],
                      open_decisions=[], control_root=tmp_path)
    assert "No class 1 or 2 deadlines on record" in r["body"]
    assert "register is incomplete" in r["body"]
    assert "O-03" in r["body"]


def test_a_populated_horizon_carries_the_same_caveat(conn, tmp_path):
    """The caveat used to fire only on an empty horizon, which had it
    backwards. A horizon showing dates is the one a reader mistakes for
    coverage — an empty one is obviously empty.
    """
    r = weekly_report(conn, as_of=AS_OF, horizon=_horizon(), open_items=[],
                      open_decisions=[], control_root=tmp_path)
    horizon_section = r["body"].split("2. OPEN")[0]
    assert "register is incomplete" in horizon_section
    assert "do not read this list as coverage" in horizon_section


def test_disputes_standing_line(conn, tmp_path):
    conn.execute(
        "INSERT INTO disputes (raised_by, raised_at, state, source) VALUES"
        " ('a.elsayed@ubcsis.com', ?, 'PENDING', 'LIVE')",
        (datetime(2026, 8, 3, 9, 0).isoformat(),),
    )
    conn.commit()
    r = weekly_report(conn, as_of=AS_OF, horizon=[], open_items=[],
                      open_decisions=[], control_root=tmp_path)
    assert "1 disputes pending adjudication, oldest 10 days" in r["body"]


def test_flags_and_register_deltas_from_db(conn, tmp_path):
    conn.execute(
        "INSERT INTO anomalies (signal, detail, source) VALUES"
        " ('S1', '[DUPLICATE_INVOICE] Steel Co: value 50000 twice', 'LIVE')")
    conn.commit()
    r = weekly_report(conn, as_of=AS_OF, horizon=[], open_items=[],
                      open_decisions=[], control_root=tmp_path)
    assert "[S1] [DUPLICATE_INVOICE]" in r["body"]
    assert "anomalies: +1 rows" in r["body"]


def test_standing_limitations_bilingual(conn, tmp_path):
    r = weekly_report(conn, as_of=AS_OF, horizon=[], open_items=[],
                      open_decisions=[], control_root=tmp_path)
    assert "not visible to this system" in r["body"]
    assert "غير مرئية لهذا النظام" in r["body"]
    assert "Their contents are not assessed" in r["body"]
    assert "ولا يتم تقييم محتواها" in r["body"]


def test_open_decisions_and_open_items(conn, tmp_path):
    items = [OpenItem("OPS-WPR-001", 3, "Weekly progress", "a.elsayed@ubcsis.com",
                      4, "L2")]
    r = weekly_report(conn, as_of=AS_OF, horizon=[], open_items=items,
                      open_decisions=["O-02: authority thresholds", "O-03: statutory calendar"],
                      control_root=tmp_path)
    assert "4 working days outstanding — stage L2" in r["body"]
    assert "OPEN: O-02" in r["body"] and "OPEN: O-03" in r["body"]


def test_nothing_outstanding_says_which_kind_of_nothing_it_is(conn, tmp_path):
    """Six tracked obligations and zero tracked obligations both used to
    print "None on record", and they are opposite facts.

    This went live the day the starter register was approved: six class 3
    obligations carrying real deadlines, and a report that said nothing
    about them. An empty section that reads as a clear week is the
    absence-looking-like-compliance failure §11 exists to prevent,
    aimed at Control's own reporting.
    """
    from dataclasses import dataclass

    @dataclass
    class Tracked:
        item_id: str
        obligation_class: int
        name: str
        owner: str
        due: date

    tracked = [
        Tracked("OPS-HRA-001", 3, "Monthly Payroll Register",
                "hr@ubcsis.com", date(2026, 9, 25)),
        Tracked("OPS-TO-001", 3, "Weekly Site Progress Report",
                "shymaa@ubcsis.com", date(2026, 8, 16)),
    ]
    r = weekly_report(conn, as_of=AS_OF, horizon=[], open_items=[],
                      open_decisions=[], control_root=tmp_path,
                      tracked=tracked)
    body = r["body"]
    assert "2 class 3 obligation(s) are tracked" in body
    assert "Weekly Site Progress Report" in body     # the nearest, not the first
    assert "shymaa@ubcsis.com" in body
    assert "16-Aug-2026 (T-3)" in body


def test_nothing_tracked_is_reported_as_an_empty_register(conn, tmp_path):
    r = weekly_report(conn, as_of=AS_OF, horizon=[], open_items=[],
                      open_decisions=[], control_root=tmp_path, tracked=[])
    assert "nothing tracked either" in r["body"]
    assert "empty register rather than a clear week" in r["body"]


def test_xlsx_export_written(conn, tmp_path):
    from pathlib import Path

    import openpyxl

    r = weekly_report(conn, as_of=AS_OF, horizon=_horizon(), open_items=[],
                      open_decisions=[], control_root=tmp_path)
    path = Path(r["xlsx_path"])
    assert path.exists()
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["Horizon", "Open Items", "Flags"]
    assert wb["Horizon"].max_row == 4  # header + 3 items (export carries all)


def test_cc_metric_line(conn, tmp_path):
    metric = {"CLOSED_OBSERVED_REPLY": 3, "CLOSED_DECLARED": 1, "BREACHED": 1,
              "OPEN": 2, "observed_share": 0.75}
    r = weekly_report(conn, as_of=AS_OF, horizon=[], open_items=[],
                      open_decisions=[], control_root=tmp_path, cc_metric=metric)
    assert "observed-reply share 75%" in r["body"]
    assert "O-05 evidence" in r["body"]
