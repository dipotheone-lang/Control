"""Weekly management report — charter §11.

Section order is chartered: the class 1 & 2 horizon is ALWAYS first.
The hard rule governs every number: if it cannot be traced to a
database row that traces to a received document, it does not appear —
the gap is stated instead. Empty sections therefore say what is missing
rather than showing a reassuring zero.

Standing lines carried verbatim in both languages (§3.1a until O-05,
§12.1.4 permanently), and the v4.3 §8.4 dispute-visibility line:
"n disputes pending, oldest x days."

Reports are ALWAYS drafts (§10) — the caller submits the result as a
MANAGEMENT_REPORT outbound, which the gate table keeps at DRAFT in
every mode.
"""

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path

from . import HaltError
from .anomaly import SUPPRESSED
from .scope import (
    MailboxScope, limitation_lines, load_scope_file, open_precondition_lines,
)

# Retained as the Option A wording. `scope.limitation_lines` selects the
# line that is true of the scope actually in force (§3.1a, D-07).
LIMITATION_SHARED_MAILBOX_EN = (
    "External SLA coverage is limited to threads copied to control@. "
    "Traffic in sales@ and procure@ is not visible to this system."
)
LIMITATION_SHARED_MAILBOX_AR = (
    "تقتصر تغطية مواعيد الرد الخارجية على المراسلات المحوّلة إلى control@. "
    "المراسلات في sales@ و procure@ غير مرئية لهذا النظام."
)
LIMITATION_CONFIDENTIAL_EN = (
    "Client-confidential documents are tracked for receipt and timeliness only. "
    "Their contents are not assessed. Accuracy, completeness, and manual "
    "conformance for these items rest with the responsible department, not "
    "with Control."
)
LIMITATION_CONFIDENTIAL_AR = (
    "يتم متابعة المستندات السرية الخاصة بالعملاء من حيث الاستلام والالتزام "
    "بالمواعيد فقط، ولا يتم تقييم محتواها. تظل مسؤولية الدقة والاكتمال "
    "ومطابقة الدليل لهذه البنود على الإدارة المختصة وليس على النظام."
)


@dataclass
class HorizonItem:
    item_id: str
    obligation_class: int      # 1 or 2
    name: str
    owner: str
    due: date
    status: str                # e.g. OPEN, ALERTED, RESOLVED


@dataclass
class OpenItem:
    item_id: str
    obligation_class: int
    name: str
    owner: str
    days_outstanding: int
    stage: str                 # e.g. L1, L2, L3, STOPPED


def _horizon_section(horizon: list[HorizonItem], as_of: date) -> list[str]:
    lines = ["1. CLASS 1 & 2 HORIZON — NEXT 30 DAYS"]
    window = [h for h in horizon
              if h.obligation_class in (1, 2)
              and as_of <= h.due <= as_of + timedelta(days=30)]
    if not window:
        lines.append(
            "   No class 1 or 2 deadlines on record for the next 30 days.")
    else:
        for h in sorted(window, key=lambda x: x.due):
            days = (h.due - as_of).days
            lines.append(
                f"   [{h.obligation_class}] {h.due:%d-%b-%Y} (T-{days}) "
                f"{h.name} — owner {h.owner} — {h.status}"
            )
    # The caveat runs in BOTH branches. It used to fire only on an empty
    # horizon, which had it backwards: a horizon showing four dates is
    # more likely to be read as coverage than one showing none, so the
    # populated case is where the reminder is worth most. This shows
    # what is on record; section 6 shows what is not.
    lines.append(
        "   This horizon is only as complete as the register behind it — "
        "the register is incomplete until the statutory calendar is "
        "advisor-verified (O-03) and the commercial registers are built. "
        "Gaps are itemised below; do not read this list as coverage."
    )
    return lines


def _open_items_section(open_items: list[OpenItem]) -> list[str]:
    lines = ["2. OPEN ITEMS BY CLASS"]
    if not open_items:
        lines.append("   None on record.")
        return lines
    for item in sorted(open_items, key=lambda x: (x.obligation_class, -x.days_outstanding)):
        lines.append(
            f"   [{item.obligation_class}] {item.name} — owner {item.owner} — "
            f"{item.days_outstanding} working days outstanding — stage {item.stage}"
        )
    return lines


def _sla_section(conn, cc_metric: dict | None) -> list[str]:
    lines = ["3. EXTERNAL SLA"]
    rows = conn.execute(
        "SELECT t.thread_id, t.category, t.owner FROM external_threads t"
        " JOIN (SELECT thread_id, MAX(id) AS mid FROM external_threads GROUP BY thread_id) m"
        " ON t.id = m.mid WHERE t.state = 'BREACHED'"
    ).fetchall()
    if rows:
        for thread_id, category, owner in rows:
            lines.append(f"   BREACHED: {thread_id} ({category}) — owner {owner} — "
                         "no reply visible to Control")
    else:
        lines.append("   No breached threads on record.")
    if cc_metric:
        share = cc_metric.get("observed_share")
        lines.append(
            "   CC-compliance (O-05 evidence): "
            f"{cc_metric.get('CLOSED_OBSERVED_REPLY', 0)} closed by observed reply, "
            f"{cc_metric.get('CLOSED_DECLARED', 0)} by CLOSED declaration, "
            f"{cc_metric.get('BREACHED', 0)} breached, {cc_metric.get('OPEN', 0)} open"
            + (f" — observed-reply share {share:.0%}" if share is not None else "")
        )
    return lines


def _register_deltas_section(conn, since: datetime) -> list[str]:
    lines = ["4. REGISTER DELTAS (LAST 7 DAYS)"]
    for table in ("submissions", "anomalies", "disputes", "external_threads"):
        n = conn.execute(
            f"SELECT COUNT(*) FROM {table} WHERE posted_at >= ?",
            (since.isoformat(sep=" "),),
        ).fetchone()[0]
        lines.append(f"   {table}: +{n} rows")
    return lines


def _flags_section(conn, since: datetime) -> list[str]:
    lines = ["5. ANOMALY FLAGS — S1–S4, FACTUAL, FOR CEO JUDGEMENT"]
    rows = conn.execute(
        "SELECT signal, detail, flagged_to FROM anomalies WHERE posted_at >= ?"
        " ORDER BY id", (since.isoformat(sep=" "),),
    ).fetchall()
    if not rows:
        lines.append("   None recorded this period.")
    else:
        raised = [r for r in rows if r[2] != SUPPRESSED]
        held = [r for r in rows if r[2] == SUPPRESSED]
        for signal, detail, _ in raised:
            lines.append(f"   [{signal}] {detail}")
        if held:
            # D-10: over-budget flags are reported as suppressed, never
            # silently dropped. A budget whose cost is invisible cannot
            # be reviewed against what it hid.
            lines.append(
                f"   {len(held)} flag(s) held back over the D-10 budget of "
                "10 a week — recorded, not dropped:")
            for signal, detail, _ in held:
                lines.append(f"     [{signal}] {detail}")
    lines += _signals_not_running()
    return lines


# §7.3 lists twelve S1 signals plus S2–S4. Most need a source the
# database does not yet hold, and a flags section that only showed what
# ran would read as "nothing found" rather than "most of this is not
# looking" (§1.1).
_SIGNAL_SOURCES = (
    ("S1 bank-detail change — the highest-priority signal in the system",
     "a supplier bank register to compare against"),
    ("S1 duplicate invoice, round-number clustering, PR/PO sequence gaps",
     "an invoice and purchase-order register"),
    ("S1 award concentration without competing quotations",
     "supplier award records with quotations attached"),
    ("S1 statistical outliers against learned baselines",
     "Stage F baselines with enough sample to be used (§7.2)"),
    ("S2 authority — approver ≠ originator, delegated limit covers value",
     "approval records, and the O-02 thresholds"),
    ("S3 implausible perfection", "several periods of the same metric"),
    ("S4 cross-source reconciliation",
     "two independent sources for the same figure"),
)


def _signals_not_running() -> list[str]:
    lines = ["", "   SIGNALS NOT RUNNING, AND WHAT EACH NEEDS:"]
    for signal, needs in _SIGNAL_SOURCES:
        lines.append(f"     {signal} — needs {needs}")
    lines.append(
        "   Out-of-hours timestamps and near-miss sender domains are the "
        "signals running today. Both work on metadata alone, so they run on "
        "confidential items too (§12.1.3).")
    return lines


# Config files that may carry an `interim:` block with a review date.
# Every interim position is registered here, so adding one cannot
# accidentally create a decision nothing ever chases.
_INTERIM_FILES = ("authority.yaml", "continuity.yaml")


def interim_reviews_due(config_dir: Path, as_of: date) -> list[str]:
    """Interim positions whose review date has arrived or is close.

    A deliberate interim decision is legitimate; one that quietly
    outlives its review date is not. These surface every week from the
    date they fall due, so an operating position cannot become permanent
    by silence.

    The wording comes from the config rather than from here, because
    the position knows what it is and this function does not.
    """
    import yaml

    due: list[str] = []
    for filename in _INTERIM_FILES:
        path = Path(config_dir) / filename
        if not path.is_file():
            continue
        try:
            data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
        except Exception:
            continue
        interim = (data.get("interim") or {}) if isinstance(data, dict) else {}
        if not interim.get("active"):
            continue
        review = interim.get("review_due")
        if not review:
            continue
        review_date = review if isinstance(review, date) else None
        if review_date is None:
            try:
                review_date = datetime.fromisoformat(str(review)).date()
            except ValueError:
                continue

        subject = interim.get("subject") or filename
        position = interim.get("position") or "interim position"
        decision = interim.get("decision") or ""
        tag = f" ({decision})" if decision else ""
        note = " ".join(str(interim.get("note") or "").split())

        days = (as_of - review_date).days
        if days >= 0:
            line = (f"{subject}: {position} is {days} day(s) past its "
                    f"{review_date:%d-%b-%Y} review{tag}.")
            due.append(f"{line} {note}".rstrip())
        elif days >= -7:
            due.append(f"{subject}: interim position reviews "
                       f"{review_date:%d-%b-%Y} ({-days} days){tag}.")
    return due


def _golden_set_notes(control_root: Path, as_of: date) -> list[str]:
    """Golden-set batches still with the CEO (§13.1).

    §13.1 states the dependency plainly: with CEO-only assignment,
    Phase 1 cannot complete without Ahmed's time, and a batch stalled
    beyond two weeks is to be raised as a deployment blocker rather than
    quietly waited out. Raising it is this line.
    """
    from .golden_worksheet import ledger_lines, load_ledger

    ledger = Path(control_root) / "tests" / "golden-set" / "worksheets" / \
        "batches.yaml"
    try:
        return ledger_lines(load_ledger(ledger), as_of)
    except Exception:
        # A malformed ledger must not take down the weekly report; it
        # surfaces on the `golden` command, which is where it is fixed.
        return []


def _transport_note(config_dir: Path) -> list[str]:
    """The interim transport route, while one is in force (D-08)."""
    import yaml

    from .transport import interim_route_note

    path = Path(config_dir) / "transport.yaml"
    if not path.is_file():
        return []
    try:
        data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    except Exception:
        return []
    note = interim_route_note(data)
    return [note] if note else []


def _continuity_notes(config_dir: Path, as_of: date) -> list[str]:
    """§13.3 continuity: backup age and last restore test.

    An unbacked-up hash chain can be truncated undetectably by a
    hardware failure, so this line appears until it has nothing to say.
    """
    import yaml

    from .backup import continuity_lines

    path = Path(config_dir) / "backup.yaml"
    if not path.is_file():
        return ["CONTINUITY: backup.yaml is missing — CONTROL_ROOT backup "
                "status is unknown (§5.2)."]
    try:
        data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    except Exception:
        return ["CONTINUITY: backup.yaml is unreadable (§5.2)."]
    return continuity_lines(data, on_date=as_of)


VALID_DISTRIBUTION_PHASES = ("PHASE_2", "STEADY_STATE")


def report_recipients(distribution: dict | None) -> tuple[list[str], str]:
    """Who receives the management report, and why that set (D-13).

    Narrowed for Phase 2 and widened at the gate. Returns the note as
    well as the list, because a recipient set that changed for a reason
    should carry the reason into the report rather than looking like a
    configuration accident.
    """
    config = (distribution or {}).get("management_reports") or {}
    phase = str(config.get("phase") or "STEADY_STATE").upper()
    if phase not in VALID_DISTRIBUTION_PHASES:
        raise HaltError(
            f"distribution.yaml: phase must be one of "
            f"{', '.join(VALID_DISTRIBUTION_PHASES)}, not {phase!r} (D-13)"
        )
    if phase == "PHASE_2":
        recipients = list(config.get("default_recipients") or [])
        note = (
            "DISTRIBUTION: narrowed to CEO and COO for Phase 2 (D-13). This "
            "report carries Control's own false positives while the system "
            "proves itself; it widens to the §11 default at the Phase 2 gate. "
            "The §12.4 usage policy is circulated to everyone regardless — "
            "the narrowing is by phase, not a private pilot."
        )
        return recipients, note
    return list(config.get("steady_state_recipients")
                or config.get("default_recipients") or []), ""


def vacancy_burden(conn, people: dict | None, since: datetime) -> list[str]:
    """§3.2 — the standing monthly line quantifying the vacancies.

        "One standing monthly line quantifying both vacancy burdens as
         hiring evidence: reporting load, transaction volume, and value
         passing through the interim arrangement."

    The charter is careful that this is evidence, not an allegation, and
    the wording here keeps that: it counts what passed through one pair
    of hands. It does not characterise anyone, and it never implies that
    anything went wrong — the point is that a structural gap nobody
    measures is a gap nobody fills.
    """
    people = people or {}
    vacant = list(people.get("vacancies") or [])
    if not vacant:
        return []

    interim = sorted({str(p.get("interim") or "") for p in vacant} - {""})
    holders = interim or ["the interim holder"]
    lines = [
        f"SOD / VACANCY BURDEN (§3.2): {len(vacant)} vacant role(s) — "
        + ", ".join(sorted(str(p.get("role", "?")) for p in vacant))
        + f". Covered on an interim basis by {', '.join(holders)}."
    ]

    counted = False
    for holder in interim:
        submissions = conn.execute(
            "SELECT COUNT(*) FROM submissions WHERE submitted_by = ?"
            " AND posted_at >= ?", (holder, since.isoformat(sep=" ")),
        ).fetchone()[0]
        threads = conn.execute(
            "SELECT COUNT(DISTINCT thread_id) FROM external_threads"
            " WHERE owner = ? AND posted_at >= ?",
            (holder, since.isoformat(sep=" ")),
        ).fetchone()[0]
        quotes = conn.execute(
            "SELECT COUNT(*), COALESCE(SUM(amount), 0), "
            " COUNT(DISTINCT currency_code) FROM registers_quotations"
            " WHERE owner = ?", (holder,),
        ).fetchone()
        if not (submissions or threads or quotes[0]):
            continue
        counted = True
        # §5.2: never total across currencies without a stated basis.
        value = (f"{quotes[1]:,.0f} EGP" if quotes[2] <= 1
                 else f"{quotes[0]} quotations across {quotes[2]} currencies "
                      "— not totalled, no stated FX basis")
        lines.append(
            f"   {holder}: {submissions} submission(s), {threads} external "
            f"thread(s) owned, {quotes[0]} quotation(s) on the register "
            f"({value})."
        )

    if not counted:
        lines.append(
            "   No transaction volume recorded yet against the interim "
            "holder(s). This is an empty register, not a light workload — "
            "the figures become evidence once Phase 2 is running (§1.1)."
        )
    lines.append(
        "   Standing recommendation (§3.2): of the two vacancies, filling "
        "procurement first restores the more valuable separation, because "
        "it splits cost from price."
    )
    return lines


def _holiday_notes(config_dir: Path, as_of: date) -> list[str]:
    """§8.3: the holiday calendar, empty or stale."""
    import yaml

    from .calendar import holiday_calendar_status

    path = Path(config_dir) / "sla.yaml"
    if not path.is_file():
        return []
    try:
        data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    except Exception:
        return ["HOLIDAY CALENDAR: sla.yaml is unreadable (§8.3)."]
    return holiday_calendar_status(data, as_of)


def _distribution_note(config_dir: Path) -> list[str]:
    import yaml

    path = Path(config_dir) / "distribution.yaml"
    if not path.is_file():
        return []
    try:
        data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    except Exception:
        return []
    _, note = report_recipients(data)
    return [note] if note else []


def _vacancy_notes(conn, config_dir: Path, since: datetime) -> list[str]:
    import yaml

    path = Path(config_dir) / "people.yaml"
    if not path.is_file():
        return []
    try:
        data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    except Exception:
        return []
    return vacancy_burden(conn, data, since)


def _decisions_section(conn, as_of: date, open_decisions: list[str]) -> list[str]:
    lines = ["6. DECISIONS REQUIRED"]
    pending = conn.execute(
        "SELECT raised_at FROM disputes d"
        " JOIN (SELECT COALESCE(submission_id, id) k, MAX(id) mid FROM disputes GROUP BY k) m"
        " ON d.id = m.mid WHERE d.state = 'PENDING'"
    ).fetchall()
    if pending:
        oldest = min(datetime.fromisoformat(r[0]).date() for r in pending)
        age = (as_of - oldest).days
        lines.append(
            f"   {len(pending)} disputes pending adjudication, oldest {age} days (§8.4)."
        )
        lines.append(
            "   Each suspends the escalation clock on its item until ruled: "
            "python -m control disputes")
    else:
        lines.append("   No disputes pending.")

    # §8.6: a repeatedly-rejected disputant is a systemic finding, raised
    # as a pattern and never re-argued item by item (§8.4).
    from .disputes import rejection_pattern
    for line in rejection_pattern(conn):
        lines.append(f"   {line}")
    for decision in open_decisions:
        lines.append(f"   OPEN: {decision}")
    return lines


def _xlsx_export(path: Path, horizon: list[HorizonItem],
                 open_items: list[OpenItem], flags: list[tuple]) -> Path:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Horizon"
    ws.append(["class", "due", "item", "owner", "status"])
    for h in sorted(horizon, key=lambda x: x.due):
        ws.append([h.obligation_class, h.due.isoformat(), h.name, h.owner, h.status])
    ws2 = wb.create_sheet("Open Items")
    ws2.append(["class", "item", "owner", "days_outstanding", "stage"])
    for item in open_items:
        ws2.append([item.obligation_class, item.name, item.owner,
                    item.days_outstanding, item.stage])
    ws3 = wb.create_sheet("Flags")
    ws3.append(["signal", "detail"])
    for signal, detail in flags:
        ws3.append([signal, detail])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def weekly_report(
    conn,
    *,
    as_of: date,
    config_dir: Path | None = None,
    horizon: list[HorizonItem],
    open_items: list[OpenItem],
    open_decisions: list[str],
    control_root: Path,
    cc_metric: dict | None = None,
) -> dict:
    since = datetime.combine(as_of - timedelta(days=7), datetime.min.time())

    sections: list[str] = []
    sections += _horizon_section(horizon, as_of)
    sections.append("")
    sections += _open_items_section(open_items)
    sections.append("")
    sections += _sla_section(conn, cc_metric)
    sections.append("")
    sections += _register_deltas_section(conn, since)
    sections.append("")
    sections += _flags_section(conn, since)
    sections.append("")
    extra: list[str] = []
    scope = MailboxScope()
    if config_dir:
        extra += interim_reviews_due(config_dir, as_of)
        scope = load_scope_file(config_dir)
        extra += open_precondition_lines(scope)
        extra += _transport_note(config_dir)
        extra += _continuity_notes(config_dir, as_of)
        extra += _holiday_notes(config_dir, as_of)
        extra += _vacancy_notes(conn, config_dir, since)
        extra += _distribution_note(config_dir)
    extra += _golden_set_notes(control_root, as_of)
    sections += _decisions_section(conn, as_of, open_decisions + extra)
    shared_en, shared_ar = limitation_lines(scope)

    flags_rows = conn.execute(
        "SELECT signal, detail FROM anomalies WHERE posted_at >= ?",
        (since.isoformat(sep=" "),),
    ).fetchall()
    xlsx_path = _xlsx_export(
        Path(control_root) / "data" / "exports" / f"weekly-{as_of.isoformat()}.xlsx",
        horizon, open_items, flags_rows,
    )

    en = "\n".join(
        [f"WEEKLY CONTROL REPORT — {as_of:%d-%b-%Y}", ""]
        + sections
        + ["", "STANDING LIMITATIONS",
           f"   {shared_en}",
           f"   {LIMITATION_CONFIDENTIAL_EN}"]
    )
    ar = "\n".join([
        f"تقرير كنترول الأسبوعي — {as_of:%d-%b-%Y}",
        "",
        "ملخص: القسم التفصيلي أعلاه باللغة الإنجليزية؛ الأرقام مطابقة.",
        f"مواعيد نهائية (فئة 1 و2) خلال 30 يوماً: {len([h for h in horizon if as_of <= h.due <= as_of + timedelta(days=30)])}",
        f"بنود مفتوحة: {len(open_items)}",
        "",
        "القيود الدائمة:",
        f"   {shared_ar}",
        f"   {LIMITATION_CONFIDENTIAL_AR}",
    ])

    return {
        "subject": f"[CONTROL] Weekly Report — {as_of:%d-%b-%Y}",
        "body": en + "\n\n" + "─" * 40 + "\n" + ar,
        "xlsx_path": str(xlsx_path),
    }
