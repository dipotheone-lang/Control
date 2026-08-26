"""Controlled forms — Stage C's form inventory (§6, check C2).

C2 asks whether a submission arrived on the current revision of the right
form, and §7.1 requires the finding to *"name old vs. current revision
explicitly"*. Neither is possible without knowing which forms exist and
which revision is current. The charter's Stage C says so directly:
identify form codes and current revisions, and where two revisions
compete, mark it `AMBIGUOUS — CEO DECISION` rather than taking the newer.

On this drive the answer is not inferred. The management system carries
its own register: a master index naming each manual and its form
workbook, and inside each workbook an `Index` sheet listing every form
code, title and type, with the revision stated in the sheet header. That
is a controlled document describing itself, which is far better evidence
than a filename heuristic, so this module reads it rather than guessing.

**Reading the register is not the same as confirming it.** The register
says a form exists; it does not say the company uses it, and Stage D's
"ghost requirement" category exists for exactly the gap between the two.
This module reports what the register contains. Whether a form is live is
a question for the obligation register and, ultimately, for the CEO.
"""

import re
from dataclasses import dataclass, field
from pathlib import Path

# `Rev. 00` in a sheet header, in either of the two shapes the workbooks
# use. The revision is a property of the register, not of one form, so a
# form with no revision of its own inherits its workbook's.
_REVISION = re.compile(r"(?i)\brev\.?\s*([0-9]+(?:\.[0-9]+)?)")
_MANUAL_CODE = re.compile(r"\b(UBC-[A-Z]{3}-MAN-\d{3})\b")
_FORM_CODE = re.compile(r"^F-[A-Z]{2,4}-\d{2,3}$")

# Words that carry no discriminating power when matching a document
# against a form title. Without this "Report" matches every form with
# "Report" in it and the best match is whichever sorted first.
_STOPWORDS = frozenset((
    "the", "a", "an", "and", "or", "of", "for", "to", "in", "on", "at",
    "by", "with", "form", "sheet", "template", "record", "co", "ub",
    "ubcsis", "united", "brothers", "no", "rev", "date", "new", "final",
))


@dataclass
class ControlledForm:
    code: str
    title: str
    kind: str
    manual: str
    manual_code: str
    revision: str
    workbook: str

    @property
    def tokens(self) -> set:
        return _tokens(self.title)


@dataclass
class FormsRegister:
    forms: list = field(default_factory=list)
    manuals: list = field(default_factory=list)
    # Workbooks found on the drive that the master index does not name,
    # and manuals the master index names but whose workbook is missing.
    # Both are C2 gaps and neither is resolvable by this module.
    unindexed_workbooks: list = field(default_factory=list)
    missing_workbooks: list = field(default_factory=list)
    revision_conflicts: dict = field(default_factory=dict)

    def by_code(self, code: str):
        for form in self.forms:
            if form.code.upper() == code.upper():
                return form
        return None


def _tokens(text: str) -> set:
    words = re.findall(r"[0-9a-zA-Z]+", str(text).lower())
    return {w for w in words if len(w) > 2 and w not in _STOPWORDS}


def read_master_index(path: Path) -> list:
    """The master index: one row per manual, naming its form workbook."""
    import openpyxl

    rows = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 1, max_col=6):
            values = [c.value for c in row]
            code = next((str(v) for v in values
                         if v and _MANUAL_CODE.fullmatch(str(v).strip())), "")
            if not code:
                continue
            name = next((str(v) for v in values
                         if v and isinstance(v, str) and "MAN-" not in v
                         and ".xlsx" not in v and len(str(v)) > 6), "")
            workbook = next((str(v) for v in values
                             if v and str(v).lower().endswith(".xlsx")), "")
            rows.append({"manual": name.strip(), "manual_code": code,
                         "workbook": workbook.strip()})
    finally:
        wb.close()
    return rows


def read_forms_workbook(path: Path) -> tuple[list, str, str]:
    """One manual's controlled forms register.

    Returns (rows, manual_code, revision). Only the `Index` sheet is
    read: the per-form sheets are the blank forms themselves, and
    treating a blank form as a submission is how a form inventory turns
    into a fabricated register of things nobody filed.
    """
    import openpyxl

    rows: list = []
    manual_code = revision = ""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if "Index" not in wb.sheetnames:
            return rows, manual_code, revision
        ws = wb["Index"]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 1, max_col=4):
            values = [c.value for c in row]
            first = str(values[0]).strip() if values[0] else ""
            if not manual_code:
                found = _MANUAL_CODE.search(first)
                if found:
                    manual_code = found.group(1)
            if not revision:
                found = _REVISION.search(first)
                if found:
                    revision = found.group(1)
            if _FORM_CODE.fullmatch(first):
                rows.append({
                    "code": first,
                    "title": str(values[1] or "").strip(),
                    "kind": str(values[2] or "").strip() or "Form",
                })
    finally:
        wb.close()
    return rows, manual_code, revision


def build(manual_root: Path) -> FormsRegister:
    """Read every forms workbook under `manual_root`.

    The master index is used to say which workbooks were expected, not to
    decide which are read: a workbook present on the drive but absent
    from the index is still read and then reported as unindexed. A form
    that exists but is not registered is a C2 problem either way, and the
    version that stays silent is the worse one.
    """
    manual_root = Path(manual_root)
    register = FormsRegister()

    expected: dict[str, dict] = {}
    for master in sorted(manual_root.rglob("*MASTER_INDEX*.xlsx")):
        for row in read_master_index(master):
            if row["workbook"]:
                expected[row["workbook"].lower()] = row

    seen: set = set()
    for workbook in sorted(manual_root.rglob("*Forms*.xlsx")):
        if "MASTER_INDEX" in workbook.name.upper():
            continue
        rows, manual_code, revision = read_forms_workbook(workbook)
        if not rows:
            continue
        seen.add(workbook.name.lower())
        declared = expected.get(workbook.name.lower())
        manual_name = (declared or {}).get("manual") or workbook.parent.name
        code = manual_code or (declared or {}).get("manual_code") or ""
        register.manuals.append({
            "manual": manual_name, "manual_code": code,
            "revision": revision or "NOT PROVIDED",
            "workbook": workbook.name, "forms": len(rows),
            "indexed": declared is not None,
        })
        for row in rows:
            register.forms.append(ControlledForm(
                code=row["code"], title=row["title"], kind=row["kind"],
                manual=manual_name, manual_code=code,
                revision=revision or "NOT PROVIDED",
                workbook=workbook.name))
        if declared is None:
            register.unindexed_workbooks.append(workbook.name)

    register.missing_workbooks = sorted(
        name for name in expected if name not in seen)

    # A form code appearing twice at different revisions is exactly the
    # AMBIGUOUS — CEO DECISION case: evaluating against the wrong
    # revision produces a confident, wrong verdict (§6 Stage C).
    by_code: dict[str, set] = {}
    for form in register.forms:
        by_code.setdefault(form.code.upper(), set()).add(form.revision)
    register.revision_conflicts = {
        code: sorted(revs) for code, revs in by_code.items() if len(revs) > 1}

    return register


# ---- matching a real document against the register --------------------

# A title must be almost entirely present in the document name before a
# form is proposed as governing it. Set from measurement, not taste: at
# 0.5 the real drive produced "Steel & Piping Progress Report for July
# 2022" -> F-PMO-05 *Initiative Progress Report* and "Insulation Report
# for November 2021" -> F-RIA-17 *Investigation Report Template*. Both
# score respectably on shared words and both are wrong, and a wrong form
# code on a register row becomes a C2 finding naming a revision the
# document was never meant to be on.
MATCH_MINIMUM = 0.75

# How far the best match must beat the runner-up. Short form titles
# overlap heavily — six of the 240 contain "Audit Report" — so a bare
# threshold still picks whichever sorted first among equals. A tie is
# `AMBIGUOUS — CEO DECISION`, which is what §6 Stage C already calls it.
MATCH_MARGIN = 0.15


@dataclass(frozen=True)
class Match:
    form: ControlledForm | None
    score: float
    runner_up: ControlledForm | None = None
    runner_up_score: float = 0.0
    ambiguous: bool = False

    @property
    def matched(self) -> bool:
        return self.form is not None and not self.ambiguous


def match(name: str, register: FormsRegister, *,
          minimum: float = MATCH_MINIMUM,
          margin: float = MATCH_MARGIN) -> Match:
    """The controlled form governing a document name, if one clearly does.

    The score is the share of the form title's *own* significant words
    present in the document name. Anchoring on the title rather than the
    document stops a long descriptive filename from matching everything
    by accident.

    Two ways to return nothing, and they are different findings. Below
    `minimum` there is no candidate at all — the series is *formless*
    against this register. Within `margin` of the runner-up there are
    two candidates and no basis to choose — that is ambiguous, and §6
    Stage C is explicit that competing forms go to the CEO rather than
    to whichever the code happened to reach first.
    """
    target = _tokens(name)
    if not target:
        return Match(None, 0.0)

    scored = []
    for form in register.forms:
        title_tokens = form.tokens
        if not title_tokens:
            continue
        scored.append((len(title_tokens & target) / len(title_tokens), form))
    if not scored:
        return Match(None, 0.0)

    scored.sort(key=lambda pair: (-pair[0], pair[1].code))
    best_score, best = scored[0]
    second_score, second = (scored[1] if len(scored) > 1 else (0.0, None))

    if best_score < minimum:
        return Match(None, best_score, second, second_score)
    if second is not None and (best_score - second_score) < margin:
        return Match(best, best_score, second, second_score, ambiguous=True)
    return Match(best, best_score, second, second_score)


def render_inventory(register: FormsRegister) -> str:
    """FORM-INVENTORY as markdown; the .xlsx export is written beside it."""
    lines = [
        "# FORM INVENTORY — Stage C",
        "",
        "Every controlled form the management system registers, read from "
        "the registers themselves rather than inferred from filenames.",
        "",
        f"**{len(register.forms)} forms across {len(register.manuals)} "
        "manuals.**",
        "",
        "Check C2 names the old revision against the current one (§7.1). "
        "That sentence is only writable because the revision below comes "
        "from the controlled register's own header.",
        "",
    ]

    if register.revision_conflicts:
        lines += [
            "## AMBIGUOUS — CEO DECISION: competing revisions", "",
            "Evaluating a submission against the wrong revision produces a "
            "confident, wrong verdict. Confirm which governs before C2 runs "
            "against these.", "",
        ]
        for code, revisions in sorted(register.revision_conflicts.items()):
            lines.append(f"- **{code}** — revisions {', '.join(revisions)}")
        lines.append("")

    if register.unindexed_workbooks or register.missing_workbooks:
        lines += ["## Register gaps", ""]
        for name in register.unindexed_workbooks:
            lines.append(
                f"- `{name}` holds controlled forms but the master index "
                "does not name it. The forms are real; the register of "
                "registers is incomplete.")
        for name in register.missing_workbooks:
            lines.append(
                f"- the master index names `{name}` and no such workbook was "
                "found. Those forms cannot be checked against.")
        lines.append("")

    lines += ["## Manuals", "",
              "| Manual | Code | Rev | Forms | In master index |",
              "|---|---|---|---|---|"]
    for entry in sorted(register.manuals, key=lambda m: m["manual"]):
        lines.append(
            f"| {entry['manual']} | {entry['manual_code'] or '—'} | "
            f"{entry['revision']} | {entry['forms']} | "
            f"{'yes' if entry['indexed'] else 'NO'} |")

    lines += ["", "## Forms", "",
              "| Code | Title | Type | Manual | Rev |", "|---|---|---|---|---|"]
    for form in sorted(register.forms, key=lambda f: f.code):
        lines.append(
            f"| {form.code} | {form.title} | {form.kind} | "
            f"{form.manual_code or form.manual} | {form.revision} |")

    lines += [
        "",
        "## What this inventory does not say",
        "",
        "- **That any of these forms is in use.** The register lists what "
        "  exists. Whether a form has ever been filed is a question for the "
        "  obligation register, where a registered form with no document "
        "  evidence is recorded as a *ghost requirement* (§6 Stage D).",
        "- **That a document matching a title is on that form.** A match on "
        "  words is a proposal for the CEO, never a C2 verdict.",
    ]
    return "\n".join(lines) + "\n"


def write_xlsx(register: FormsRegister, path: Path) -> Path:
    """Stage J deliverable 3 — FORM-INVENTORY.xlsx."""
    import openpyxl

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Forms"
    ws.append(["Form code", "Title", "Type", "Manual", "Manual code",
               "Revision", "Workbook"])
    for form in sorted(register.forms, key=lambda f: f.code):
        ws.append([form.code, form.title, form.kind, form.manual,
                   form.manual_code, form.revision, form.workbook])

    ws2 = wb.create_sheet("Manuals")
    ws2.append(["Manual", "Manual code", "Revision", "Forms",
                "In master index"])
    for entry in sorted(register.manuals, key=lambda m: m["manual"]):
        ws2.append([entry["manual"], entry["manual_code"], entry["revision"],
                    entry["forms"], "yes" if entry["indexed"] else "NO"])

    ws3 = wb.create_sheet("Gaps")
    ws3.append(["Kind", "Item", "Why it matters"])
    for code, revisions in sorted(register.revision_conflicts.items()):
        ws3.append(["COMPETING REVISION", f"{code}: {', '.join(revisions)}",
                    "C2 against the wrong revision is a confident wrong "
                    "verdict (§6 Stage C)"])
    for name in register.unindexed_workbooks:
        ws3.append(["UNINDEXED WORKBOOK", name,
                    "Holds controlled forms the master index does not name"])
    for name in register.missing_workbooks:
        ws3.append(["MISSING WORKBOOK", name,
                    "Named by the master index but not found on the drive"])

    wb.save(path)
    wb.close()
    return path
